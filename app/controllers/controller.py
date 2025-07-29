import os
import re
import pandas as pd

from datetime import datetime
from app.utils.users import get_users
from app.config.configuration import get_config
from openpyxl.styles import NamedStyle


def get_file(data):
    """Obtiene el archivo Excel del servidor y lo guarda localmente."""
    # Obtener el usuario del commit y construir la ruta del archivo Excel
    username = data["head_commit"]["committer"]["username"]
    # Obtener la ruta del usuario desde la configuración
    user_path = get_users()[username]

    # Obtener la configuración del servidor y la ruta del archivo
    excel_route, sub_path, server_user, password = get_config().values()

    # Construir la ruta del archivo en el servidor y el nombre del archivo Excel
    file_path = os.path.join(sub_path, user_path)
    excel_file = f"2025 - Bitácora de desarrollo - {user_path}.xlsx"
    local_file = f"app/2025 - Bitácora de desarrollo - {user_path}.xlsx"

    # Comando para descargar el archivo Excel del servidor
    command = (
        f'smbclient "{excel_route}" -U "{server_user}%{password}" '
        f'-c "cd \\"{file_path}\\"; get \\"{excel_file}\\" \\"{local_file}\\""'
    )

    # Ejecutar el comando para descargar el archivo
    os.system(command)

    return local_file, file_path, excel_file


def upload_file(local_file, file_path, excel_file):
    """Sube el archivo Excel modificado al servidor."""
    excel_route, sub_path, server_user, password = get_config().values()

    command_upload = (
        f'smbclient "{excel_route}" -U "{server_user}%{password}" '
        f'-c "cd \\"{file_path}\\"; put \\"{local_file}\\" \\"{excel_file}\\""'
    )
    # print("Subiendo el archivo modificado...", command_upload)
    os.system(command_upload)
    return


def obtener_campos_commit(data):
    """Genera los campos necesarios para el commit."""
    mensaje = data["head_commit"]["message"]
    # Regex para extraer el ticket y la descripción del commit
    commit_regex = re.search(r"^(\[[\d-]*\]\s[A-Za-z\.\d]*)\s-\s([\w\W]+)$", mensaje)

    # Verificar si el mensaje del commit sigue el formato esperado
    if not commit_regex:
        raise ValueError("El mensaje del commit no sigue el formato esperado.")
    # Extraer el ticket y la descripción del commit
    ticket = commit_regex.group(1)
    descripcion_split = (commit_regex.group(2)).split("-")

    campo_tarea = "-".join(descripcion_split[:-1])
    campo_accion_tomada = descripcion_split[-1].strip()

    return ticket, campo_tarea, campo_accion_tomada


def calcular_campos_df(df, ticket, data):
    """Calcula los campos necesarios para el DataFrame."""
    ticket_exists = df[df["Ticket - Proyectos"] == ticket]

    # Verificar si el ticket ya existe en el DataFrame
    if ticket_exists.empty:
        hash = max(df["Hash"].values) + 1 if not df.empty else 1
        fecha_asignacion = datetime.strptime(ticket[1:11], "%Y-%m-%d").strftime(
            "%d/%m/%Y"
        )
    else:
        hash = ticket_exists["Hash"].values[0]
        fecha_asignacion = ticket_exists["Fecha de asignación"].values[-1]

    # Extraer la fecha de resolución del commit
    fecha_resolucion = datetime.strptime(
        data["head_commit"]["timestamp"], "%Y-%m-%dT%H:%M:%S%z"
    ).strftime("%d/%m/%Y")

    estado = "Terminado"

    return hash, fecha_asignacion, fecha_resolucion, estado


def aplicar_formato_fechas_excel(writer, df, sheet_name):
    """Aplica formato de fecha a las columnas de fecha en Excel."""
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]

    # Crear un estilo de fecha si no existe
    date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")
    if "date_style" not in workbook.named_styles:
        workbook.add_named_style(date_style)

    # Aplicar el formato de fecha a las columnas correspondientes
    fecha_asignacion_col = None
    fecha_resolucion_col = None

    # Encontrar las columnas de fecha
    for col_num, col_name in enumerate(df.columns, 1):
        if col_name == "Fecha de asignación":
            fecha_asignacion_col = col_num
        elif col_name == "Fecha de Resolución":
            fecha_resolucion_col = col_num

    # Aplicar formato a las celdas de fecha
    if fecha_asignacion_col:
        for row in range(
            2, len(df) + 2
        ):  # Empezar desde la fila 2 (después del header)
            cell = worksheet.cell(row=row, column=fecha_asignacion_col)
            cell.style = date_style

    if fecha_resolucion_col:
        for row in range(
            2, len(df) + 2
        ):  # Empezar desde la fila 2 (después del header)
            cell = worksheet.cell(row=row, column=fecha_resolucion_col)
            cell.style = date_style


def guardar_en_excel(data):
    """Guarda los datos del commit en un archivo Excel."""
    # Obtener el archivo Excel local, su ruta y el nombre del archivo
    local_file, file_path, excel_file = get_file(data)

    # Extraer los campos necesarios del mensaje del commit
    ticket, campo_tarea, campo_accion_tomada = obtener_campos_commit(data)

    # Cargar el archivo Excel
    df = pd.read_excel(local_file, sheet_name=1, engine="openpyxl")

    # Calcular los campos necesarios para el DataFrame
    hash, fecha_asignacion, fecha_resolucion, estado = calcular_campos_df(
        df, ticket, data
    )

    # Crear un diccionario con los datos de la nueva fila
    nueva_fila = {
        "Ticket - Proyectos": ticket,
        "Hash": hash,
        "Tarea": campo_tarea,
        "Fecha de asignación": fecha_asignacion,
        "Fecha de Resolución": fecha_resolucion,
        "Acción Tomada": campo_accion_tomada,
        "Estado": estado,
        "Notas": "",
    }
    
    # Añadir la nueva fila al DataFrame
    df = pd.concat([df, pd.DataFrame([nueva_fila])], ignore_index=True)

    # Convertir las columnas de fecha al formato adecuado
    df["Fecha de asignación"] = pd.to_datetime(
        df["Fecha de asignación"], format="%d/%m/%Y", errors="coerce"
    ).dt.date
    df["Fecha de Resolución"] = pd.to_datetime(
        df["Fecha de Resolución"], format="%d/%m/%Y"
    ).dt.date

    # Guardar el DataFrame modificado en el archivo Excel
    with pd.ExcelWriter(
        local_file, engine="openpyxl", mode="a", if_sheet_exists="replace"
    ) as writer:
        sheet_name = writer.book.sheetnames[1]
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Aplicar formato de fecha a las columnas correspondientes
        aplicar_formato_fechas_excel(writer, df, sheet_name)

    # Subir el archivo modificado al servidor
    upload_file(local_file, file_path, excel_file)

    # Eliminar archivo local después de subirlo
    os.remove(local_file)

    return
